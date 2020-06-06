#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.colors import Color
from openpyxl.utils.cell import get_column_letter
from PIL import Image
from tqdm import tqdm


LIMIT_OF_EXCEL = 0x18001


def ctuple2cstr(tup, alpha=0xff):
    r, g, b = tup
    return ("%02x%02x%02x" % (r, g, b)).upper()


def resize_tuple(width, height):
    all_pixel = width * height
    if all_pixel < LIMIT_OF_EXCEL:
        return (width, height)
    alpha = (LIMIT_OF_EXCEL / all_pixel) ** 0.5
    print("alpha = %f" % alpha)
    r_width, r_height = int(width * alpha), int(height * alpha)
    actual_pixel = r_width * r_height
    print("actual pixel to write = %d" % actual_pixel)

    return (r_width, r_height)


def main(arguments):
    inputs = arguments["<FILE>"]
    output = arguments["<EXCELFILE>"]

    print("inputs: %s" % inputs)
    print("output: %s" % output)

    wb = Workbook()
    height_in_points = 10.0
    width_in_charwidth = 1.0

    wss = []
    for fidx, f in enumerate(inputs):
        if fidx == 0:
            ws = wb.active
            ws.title = os.path.basename(f)
        else:
            ws = wb.create_sheet(os.path.basename(f))

        with Image.open(f) as im:
            width, height = im.size
            resized = resize_tuple(width, height)
            c_width, c_height = resized
            im2 = im.resize(resized)
            print("%s: format=%s, size=%s" % (f, im.format, im.size))
            if c_width != width or c_height != height:
                print("resized to: size=%s" % str(im2.size))
            bands = im2.getbands()
            cdict = {}
            for ix, b in enumerate(bands):
                if b == "R":
                    cdict["R"] = ix
                elif b == "G":
                    cdict["G"] = ix
                elif b == "B":
                    cdict["B"] = ix
                elif b == "L":
                    cdict["L"] = ix
                elif b == "A":
                    cdict["A"] = ix
            try:
                rgb = zip(list(im2.getdata(band=cdict["R"])), list(
                    im2.getdata(band=cdict["G"])), list(im2.getdata(band=cdict["B"])))
            except:
                rgb = zip(list(im2.getdata(band="L")), list(
                    im2.getdata(band="L")), list(im2.getdata(band="L")))

            rgb = list(rgb)

            bytes_written = 0
            for x in tqdm(range(c_width)):
                for y in range(c_height):
                    _ = ws.cell(column=(x + 1), row=(y + 1), value=" ")
                    bytes_written += 1
                    if bytes_written > LIMIT_OF_EXCEL:
                        break
                    color = ctuple2cstr(rgb[y * c_width + x])
                    c = ws[get_column_letter(x + 1) + ("%d" % (y + 1))]
                    c.fill = PatternFill(fgColor=color, fill_type="solid")

            for y in range(c_height):
                ws.row_dimensions[y + 1].height = height_in_points

            for x in range(c_width):
                ws.column_dimensions[get_column_letter(
                    x + 1)].width = width_in_charwidth

    wb.save(output)
